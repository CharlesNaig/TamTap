#!/usr/bin/env python3
"""
TAMTAP - RFID to Excel Registration Tool
Reads NFC cards and stores UIDs in an Excel file with auto-incrementing IDs.
Max entries: 50
"""

import sys
import time
import logging
from pathlib import Path

# Configure logging (no print debugging per contract)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Excel file path
EXCEL_FILE = Path(__file__).parent / "rfid_registry.xlsx"
MAX_ENTRIES = 50

try:
    from mfrc522 import SimpleMFRC522
    import RPi.GPIO as GPIO
    HARDWARE_AVAILABLE = True
except ImportError:
    HARDWARE_AVAILABLE = False
    logger.warning("Hardware libraries not available - running in simulation mode")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    logger.error("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def create_excel_file():
    """Create a new Excel file with headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RFID Registry"
    
    # Headers
    ws['A1'] = "tamtap_Id"
    ws['B1'] = "nfc_Id"
    
    # Style headers
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ['A1', 'B1']:
        ws[cell].font = header_font
        ws[cell].alignment = header_alignment
        ws[cell].border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    
    wb.save(EXCEL_FILE)
    logger.info(f"Created new Excel file: {EXCEL_FILE}")
    return wb


def load_or_create_excel():
    """Load existing Excel file or create new one."""
    if EXCEL_FILE.exists():
        try:
            wb = load_workbook(EXCEL_FILE)
            logger.info(f"Loaded existing Excel file: {EXCEL_FILE}")
            return wb
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return create_excel_file()
    else:
        return create_excel_file()


def get_existing_nfc_ids(ws):
    """Get all existing NFC IDs from the worksheet."""
    existing_ids = set()
    for row in range(2, ws.max_row + 1):
        nfc_id = ws.cell(row=row, column=2).value
        if nfc_id:
            existing_ids.add(str(nfc_id))
    return existing_ids


def get_next_tamtap_id(ws):
    """Get the next tamtap_Id (auto-increment from 001)."""
    max_id = 0
    for row in range(2, ws.max_row + 1):
        tamtap_id = ws.cell(row=row, column=1).value
        if tamtap_id:
            try:
                # Extract number from format like "001"
                num = int(str(tamtap_id))
                if num > max_id:
                    max_id = num
            except ValueError:
                continue
    return max_id + 1


def get_entry_count(ws):
    """Count current entries in the worksheet."""
    count = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value:
            count += 1
    return count


def add_nfc_entry(wb, ws, nfc_id):
    """Add a new NFC entry to the Excel file."""
    nfc_id_str = str(nfc_id)
    
    # Check for duplicates
    existing_ids = get_existing_nfc_ids(ws)
    if nfc_id_str in existing_ids:
        logger.error(f"ERROR: NFC ID {nfc_id_str} already exists in registry!")
        return False, "DUPLICATE"
    
    # Check max entries
    entry_count = get_entry_count(ws)
    if entry_count >= MAX_ENTRIES:
        logger.error(f"ERROR: Maximum entries ({MAX_ENTRIES}) reached!")
        return False, "MAX_REACHED"
    
    # Get next ID
    next_id = get_next_tamtap_id(ws)
    tamtap_id = f"{next_id:03d}"  # Format as 001, 002, etc.
    
    # Find next empty row
    next_row = ws.max_row + 1
    if ws.max_row == 1 and ws.cell(row=2, column=1).value is None:
        next_row = 2
    
    # Add entry with styling
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center')
    
    ws.cell(row=next_row, column=1, value=tamtap_id)
    ws.cell(row=next_row, column=2, value=nfc_id_str)
    
    for col in [1, 2]:
        ws.cell(row=next_row, column=col).border = thin_border
        ws.cell(row=next_row, column=col).alignment = center_alignment
    
    # Save immediately
    wb.save(EXCEL_FILE)
    logger.info(f"SUCCESS: Added entry - tamtap_Id: {tamtap_id}, nfc_Id: {nfc_id_str}")
    return True, tamtap_id


def read_rfid():
    """Read RFID card and return UID."""
    if HARDWARE_AVAILABLE:
        try:
            reader = SimpleMFRC522()
            uid, _ = reader.read_no_block()
            return uid
        except Exception as e:
            logger.error(f"RFID read error: {e}")
            return None
    else:
        # Simulation mode for testing
        return None


def display_current_entries(ws):
    """Display all current entries in the registry."""
    print("\n" + "=" * 40)
    print("       CURRENT RFID REGISTRY")
    print("=" * 40)
    print(f"{'tamtap_Id':<12} | {'nfc_Id':<20}")
    print("-" * 40)
    
    entry_count = 0
    for row in range(2, ws.max_row + 1):
        tamtap_id = ws.cell(row=row, column=1).value
        nfc_id = ws.cell(row=row, column=2).value
        if tamtap_id and nfc_id:
            print(f"{tamtap_id:<12} | {nfc_id:<20}")
            entry_count += 1
    
    print("-" * 40)
    print(f"Total entries: {entry_count}/{MAX_ENTRIES}")
    print("=" * 40 + "\n")


def main():
    """Main function - RFID to Excel registration loop."""
    print("\n" + "=" * 50)
    print("   TAMTAP - RFID to Excel Registration Tool")
    print("=" * 50)
    print(f"Max entries: {MAX_ENTRIES}")
    print(f"Excel file: {EXCEL_FILE}")
    print("-" * 50)
    
    # Load or create Excel file
    wb = load_or_create_excel()
    ws = wb.active
    
    # Display current entries
    display_current_entries(ws)
    
    if not HARDWARE_AVAILABLE:
        print("\n[SIMULATION MODE - No hardware detected]")
        print("Enter NFC IDs manually (or 'q' to quit, 'list' to show entries):\n")
        
        while True:
            try:
                user_input = input("Enter NFC ID: ").strip()
                
                if user_input.lower() == 'q':
                    print("Exiting...")
                    break
                
                if user_input.lower() == 'list':
                    wb = load_or_create_excel()
                    ws = wb.active
                    display_current_entries(ws)
                    continue
                
                if not user_input:
                    print("Please enter a valid NFC ID.")
                    continue
                
                # Reload to ensure latest data
                wb = load_or_create_excel()
                ws = wb.active
                
                success, result = add_nfc_entry(wb, ws, user_input)
                
                if success:
                    print(f"✓ Registered: tamtap_Id={result}, nfc_Id={user_input}")
                elif result == "DUPLICATE":
                    print(f"✗ ERROR: NFC ID '{user_input}' already exists!")
                elif result == "MAX_REACHED":
                    print(f"✗ ERROR: Maximum entries ({MAX_ENTRIES}) reached!")
                
                print()
                
            except KeyboardInterrupt:
                print("\n\nExiting...")
                break
            except Exception as e:
                logger.error(f"Error: {e}")
    else:
        print("\n[HARDWARE MODE - RC522 RFID Reader]")
        print("Place RFID card on reader (Ctrl+C to quit):\n")
        
        try:
            reader = SimpleMFRC522()
            last_read_time = 0
            debounce_delay = 2  # seconds between reads
            
            while True:
                try:
                    current_time = time.time()
                    
                    print("Waiting for card...", end='\r')
                    uid, text = reader.read()
                    
                    # Debounce to prevent multiple reads
                    if current_time - last_read_time < debounce_delay:
                        continue
                    
                    last_read_time = current_time
                    
                    if uid:
                        print(f"\nCard detected! UID: {uid}")
                        
                        # Reload Excel to ensure latest data
                        wb = load_or_create_excel()
                        ws = wb.active
                        
                        success, result = add_nfc_entry(wb, ws, uid)
                        
                        if success:
                            print(f"✓ Registered: tamtap_Id={result}, nfc_Id={uid}")
                        elif result == "DUPLICATE":
                            print(f"✗ ERROR: NFC ID '{uid}' already exists!")
                        elif result == "MAX_REACHED":
                            print(f"✗ ERROR: Maximum entries ({MAX_ENTRIES}) reached!")
                        
                        print()
                        time.sleep(1)  # Brief pause after read
                        
                except Exception as e:
                    logger.error(f"Read error: {e}")
                    time.sleep(0.5)
                    
        except KeyboardInterrupt:
            print("\n\nExiting...")
        finally:
            GPIO.cleanup()
            logger.info("GPIO cleaned up")
    
    print("\nRegistry saved to:", EXCEL_FILE)
    display_current_entries(ws)


if __name__ == "__main__":
    main()
